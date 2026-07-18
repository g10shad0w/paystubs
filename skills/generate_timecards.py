#!/usr/bin/env python3
"""Generate a multi-sheet Excel time-card workbook from a W2 paystubs PDF.

One employee per sheet. Daily hours are synthesized to sum exactly to each
employee's paystub total (a model/mock, not a real attendance record).

Days off can be specified per employee via --daysoff (a JSON file mapping a
name/alias to a list of weekday abbreviations, e.g. {"Qadir": ["Tue","Fri"]}).
Aliases are fuzzy-matched to the paystub names; any names left unmatched on both
sides are paired by order. Employees with no entry fall back to auto scheduling.

Usage:
    python3 generate_timecards.py --pdf paystubs.pdf --out timecards.xlsx \
        [--daysoff daysoff.json] [--seed 42]
"""
import argparse, json, random, re, sys, datetime, difflib
import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DAYS = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
ABBR = {"mon":0,"tue":1,"wed":2,"thu":3,"fri":4,"sat":5,"sun":6}

def hm_to_min(s):
    h, m = s.split(":"); return int(h)*60 + int(m)
def min_to_hm(t):
    return f"{t//60}:{t%60:02d}"

def parse_days(vals):
    out = []
    for v in vals:
        k = str(v).strip().lower()[:3]
        if k in ABBR: out.append(ABBR[k])
    return sorted(set(out))

def parse_pdf(path):
    """Return (employees, week_start_date). employees: list of (name, 'HH:MM')."""
    emps, period_start = [], None
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            m = re.search(r"Pay Period:\s*(\d{2}/\d{2}/\d{4})", text)
            if m and not period_start:
                period_start = datetime.datetime.strptime(m.group(1), "%m/%d/%Y").date()
            name = None
            ssn = re.search(r"^([A-Z][A-Z .'-]+?),.*\*\*\*-\*\*-", text, re.M)
            if ssn:
                name = ssn.group(1).strip()
            hrs = re.search(r"Hourly\s+(\d{1,3}:\d{2})", text)
            if name and hrs:
                emps.append((name.title(), hrs.group(1)))
    return emps, period_start

def _norm(s):
    return re.sub(r"[^a-z]", "", s.lower())

def resolve_daysoff(emps, spec):
    """Map spec {alias: [day-abbrevs]} to {full_employee_name: [day-indices]}.

    Fuzzy-match each alias to the closest paystub name; leftovers on both sides
    are paired by order (handles aliases that don't resemble the legal name).
    Returns (mapping, report) where report lists how each alias resolved.
    """
    if not spec: return {}, []
    names = [n for n, _ in emps]
    remaining = set(names)
    resolved, report = {}, []
    # score all alias/name pairs, assign greedily by best similarity
    scored = []
    for alias in spec:
        na = _norm(alias)
        for name in names:
            nn = _norm(name)
            r = difflib.SequenceMatcher(None, na, nn).ratio()
            if na and (na in nn or nn.startswith(na)):
                r = max(r, 0.9)
            scored.append((r, alias, name))
    scored.sort(reverse=True)
    used_alias = set()
    for r, alias, name in scored:
        if alias in used_alias or name not in remaining or r < 0.5:
            continue
        resolved[name] = parse_days(spec[alias]); remaining.discard(name)
        used_alias.add(alias); report.append((alias, name, r))
    # pair leftovers by order
    left_alias = [a for a in spec if a not in used_alias]
    left_names = [n for n in names if n in remaining]
    for alias, name in zip(left_alias, left_names):
        resolved[name] = parse_days(spec[alias])
        report.append((alias, name, 0.0))
    return resolved, report

def distribute(total_min, ndays, rng):
    base = total_min // ndays
    vals = [base]*ndays
    for i in range(total_min - base*ndays):
        vals[i] += 1
    for _ in range(ndays):
        i, j = rng.sample(range(ndays), 2)
        d = rng.randint(1, 4)
        vals[i] += d; vals[j] -= d
    return vals

def build(emps, week_start, out_path, daysoff=None, seed=42):
    daysoff = daysoff or {}
    rng = random.Random(seed)
    dates = [(week_start + datetime.timedelta(days=i)) for i in range(7)] if week_start else [None]*7
    date_strs = [f"{d.month}/{d.day}/{d.year}" if d else "" for d in dates]
    week_of = f"{week_start.month:02d}/{week_start.day:02d}/{week_start.year}" if week_start else ""

    navy, blue, green = "1F3864", "2E75B6", "375623"
    lightblue, lightgreen, gray = "DDEBF7", "E2EFDA", "F2F2F2"
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    wb = Workbook(); wb.remove(wb.active)
    for name, tot in emps:
        total_min = hm_to_min(tot)
        if name in daysoff:
            off_idx = set(daysoff[name])
        else:
            ndays = max(4, min(7, round(total_min/60/5.7)))
            off = 7 - ndays
            off_idx = set([1,3,5,6,4,2,0][:off]) if off > 0 else set()
        work_idx = [i for i in range(7) if i not in off_idx]
        if not work_idx:  # safety: never zero working days
            work_idx = [0]; off_idx.discard(0)
        shifts = distribute(total_min, len(work_idx), rng)
        day_hours = {i: shifts[k] for k, i in enumerate(work_idx)}

        ws = wb.create_sheet(title=name[:31])
        for c, w in enumerate([12,13,11,12,11,16], start=1):
            ws.column_dimensions[get_column_letter(c)].width = w

        ws.merge_cells("A1:F1")
        t = ws["A1"]; t.value = "EMPLOYEE TIME CARD"
        t.font = Font(bold=True, color="FFFFFF", size=20)
        t.fill = PatternFill("solid", fgColor=navy); t.alignment = center
        ws.row_dimensions[1].height = 34

        ws.merge_cells("A2:C2"); ws.merge_cells("D2:F2")
        e = ws["A2"]; e.value = f"Employee:  {name}"
        e.font = Font(bold=True, size=12); e.fill = PatternFill("solid", fgColor=lightblue); e.alignment = center
        wk = ws["D2"]; wk.value = f"Week of:  {week_of}"
        wk.font = Font(bold=True, size=12); wk.fill = PatternFill("solid", fgColor=lightblue); wk.alignment = center
        ws.row_dimensions[2].height = 24
        for col in "ABCDEF": ws[f"{col}2"].border = border

        for c, h in enumerate(["Day","Date","Clock In","Clock Out","Hours","Notes"], start=1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill("solid", fgColor=blue); cell.alignment = center; cell.border = border
        ws.row_dimensions[3].height = 20

        r = 4
        for i in range(7):
            fill = gray if i % 2 == 1 else "FFFFFF"
            for c in range(1, 7):
                cell = ws.cell(row=r, column=c); cell.border = border
                cell.fill = PatternFill("solid", fgColor=fill); cell.alignment = center
            ws.cell(row=r, column=1, value=DAYS[i]).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=r, column=2, value=date_strs[i])
            if i in off_idx:
                oc = ws.cell(row=r, column=3, value="OFF"); oc.font = Font(italic=True, color="808080")
                ws.cell(row=r, column=5, value="-")
            else:
                cin = 12*60 + rng.randint(0, 7); mins = day_hours[i]
                ws.cell(row=r, column=3, value=min_to_hm(cin))
                ws.cell(row=r, column=4, value=min_to_hm(cin+mins))
                ws.cell(row=r, column=5, value=min_to_hm(mins))
            ws.row_dimensions[r].height = 20
            r += 1

        ws.merge_cells(f"A{r}:D{r}")
        tc = ws.cell(row=r, column=1, value="TOTAL HOURS")
        tc.font = Font(bold=True, color="FFFFFF", size=12); tc.alignment = center
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=green)
            ws.cell(row=r, column=c).border = border
        hc = ws.cell(row=r, column=5, value=tot)
        hc.font = Font(bold=True, size=13); hc.alignment = center
        hc.fill = PatternFill("solid", fgColor=lightgreen); hc.border = border
        ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor=lightgreen)
        ws.cell(row=r, column=6).border = border
        ws.row_dimensions[r].height = 26
        r += 2

        ws.merge_cells(f"A{r}:C{r}"); ws.merge_cells(f"D{r}:F{r}")
        ws.cell(row=r, column=1, value="Employee Signature: ______________________").font = Font(size=10)
        ws.cell(row=r, column=4, value="Manager Signature: ______________________").font = Font(size=10)

        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_area = f"A1:F{r}"

    wb.save(out_path)
    return out_path

def verify(path, daysoff=None):
    daysoff = daysoff or {}
    wb = load_workbook(path); ok = True
    print(f"{'Sheet':<26} {'sum':>6} {'total':>6} {'off':>10}  status")
    for ws in wb.worksheets:
        s = 0; offs = []
        for i, row in enumerate(ws.iter_rows(min_row=4, max_row=10, min_col=3, max_col=5)):
            cin, _, hrs = row[0].value, row[1].value, row[2].value
            if cin == "OFF": offs.append(DAYS[i][:3])
            if hrs and hrs != "-": s += hm_to_min(hrs)
        total = None
        for row in ws.iter_rows():
            for c in row:
                if c.value == "TOTAL HOURS":
                    total = ws.cell(row=c.row, column=5).value
        got = min_to_hm(s)
        exp_off = sorted(daysoff.get(ws.title, []))
        off_ok = (exp_off == sorted(ABBR[o.lower()] for o in offs)) if ws.title in daysoff else True
        status = "OK" if (got == total and off_ok) else "MISMATCH"
        if status != "OK": ok = False
        print(f"{ws.title:<26} {got:>6} {str(total):>6} {','.join(offs):>10}  {status}")
    return ok

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdf", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--daysoff", help="JSON file: {alias: [weekday abbrevs]}")
    ap.add_argument("--seed", type=int, default=42)
    a = ap.parse_args()
    emps, week_start = parse_pdf(a.pdf)
    if not emps:
        print("No employees found in PDF.", file=sys.stderr); sys.exit(1)
    print(f"Found {len(emps)} employees; week starting {week_start}")
    daysoff = {}
    if a.daysoff:
        spec = json.load(open(a.daysoff))
        daysoff, report = resolve_daysoff(emps, spec)
        print("Days-off mapping:")
        for alias, name, r in report:
            tag = "by-order" if r == 0.0 else f"{r:.2f}"
            print(f"  {alias:<18} -> {name:<26} ({tag}) off={[DAYS[i][:3] for i in daysoff[name]]}")
    build(emps, week_start, a.out, daysoff, a.seed)
    if not verify(a.out, daysoff):
        print("Verification failed.", file=sys.stderr); sys.exit(2)
    print(f"Saved {a.out}")

if __name__ == "__main__":
    main()
